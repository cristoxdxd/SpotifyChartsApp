from io import BytesIO
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment, PatternFill, Border
from openpyxl.drawing.image import Image
import pandas as pd
from .data import top_10_year, top_5_artists_year

FONT = 'Arial'

class PDFReport(FPDF):
    def __init__(self):
        super().__init__()
        self.set_margins(10, 10, 10)
        self.background = None
        self.year = None
    
    def header(self, title=''):
        self.set_font(FONT, 'B', 24)
        self.cell(0, 10, title, 0, 1, 'L')
        self.ln(10)

    def footer(self):
        self.set_y(-15)
        self.set_font(FONT, 'I', 8)
        self.cell(0, 10, '%s' % self.page_no(), 0, 0, 'R')

    def set_background(self, image_path):
        self.background = image_path

    def set_year(self, year):
        self.year = year

    def first_page(self, year=''):
        self.set_background('public/PDF_background.png')
        self.add_page()
        if self.background:
            self.image(self.background, 0, 0, self.w, self.h)
        self.set_y(115)
        self.set_font(FONT, 'B', 85)
        self.cell(0, 5, '      Spotify', 0, 1, 'C')
        if year:
            self.set_y(150)
            self.set_font(FONT, '', 30)
            self.cell(0, 5, '      ANNUAL REPORT', 0, 1, 'C')
            self.set_y(260)
            self.set_font(FONT, 'I', 35)
            self.cell(0, 5, '         YEAR', 0, 1, 'L')
            self.set_y(260)
            self.cell(0, 5, year + '       ', 0, 1, 'R')
        else:
            self.set_y(150)
            self.set_font(FONT, '', 30)
            self.cell(0, 5, '  FULL REPORT', 0, 1, 'C')

def generate_pdf_report(year=''):
    pdf = PDFReport()
    pdf.first_page(year)

    pdf.set_year(year)
    pdf.add_page()

    if year:
        pdf.set_font(FONT, 'B', 16)
        pdf.header(f'Top 10 Songs of the Year {year}')
        for index in range(10):
            pdf.image(f"images/top10_{year}_{index}.png", w=200)
        pdf.add_page()
        pdf.header(f'Top 5 Artists of the Year {year}')
        pdf.image(f"images/top5_{year}.png", w=170)
    else:
        pdf.set_font(FONT, 'B', 16)
        pdf.header('Summary Tracks by years')
        pdf.image("images/summary.png", w=170)
        pdf.add_page()
        pdf.header('Top 10 Songs of All Time')
        for index in range(10):
            pdf.image(f"images/top10_{index}.png", w=200)
        pdf.add_page()
        pdf.header('Top 5 Artists of All Time')
        pdf.image("images/top5.png", w=170)
    
    return pdf.output(dest='S').encode('latin-1')

def generate_excel_report(year):
    FONT_NAME = 'Amercian Typewriter'
    
    wb = Workbook()
    ws = wb.active

    
    ws['E1'] = f"Spotify Dashboard {year}"

    ws['E1'].font = Font(name=FONT_NAME, size=20, bold=True, italic=True, color='FFFFFF')

    ws.row_dimensions[1].height = 30

    ws['E1'].alignment = Alignment(horizontal='center', vertical='center')

    ws['E1'].fill = PatternFill(start_color='050505',end_color='050505',fill_type='solid')

    ws.merge_cells('E1:I1')

    # img = []
    # for index in range(10):
    #     img = Image(f'images/top10_{year}_{index}.png')
    #     ws.add_image(img,f'A{str(index*6+4)}')

    img = Image(f'images/top5_{year}.png')
    ws.add_image(img,'A22')

    top_10_year_data = top_10_year(year)
    
    ws['A3'] = 'Top 10 Songs'
    ws['A3'].font = Font(name=FONT_NAME, size=16, bold=True, italic=True, color='FFFFFF')
    ws.row_dimensions[50].height = 30
    ws.column_dimensions['A'].width = 30
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A3'].fill = PatternFill(start_color='050505',end_color='050505',fill_type='solid')

    for index, row in top_10_year_data.iterrows():
        ws[f'A{index+4}'] = row['song']
        ws[f'B{index+4}'] = row['artist']
        ws[f'C{index+4}'] = row['popularity']
        ws[f'D{index+4}'] = row['danceability']
        ws[f'E{index+4}'] = row['energy']
        ws[f'F{index+4}'] = row['loudness']
        ws[f'G{index+4}'] = row['liveness']
        ws[f'H{index+4}'] = row['tempo']

    top_5_artists_year_data = top_5_artists_year(year)

    ws['A15'] = 'Top 5 Artists'
    ws['A15'].font = Font(name=FONT_NAME, size=16, bold=True, italic=True, color='FFFFFF')
    ws.row_dimensions[70].height = 30
    ws['A15'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A15'].fill = PatternFill(start_color='050505',end_color='050505',fill_type='solid')

    for index, row in top_5_artists_year_data.iterrows():
        ws[f'A{index+16}'] = row['artist']
        ws[f'B{index+16}'] = row['popularity']
        ws[f'C{index+16}'] = row['danceability']
        ws[f'D{index+16}'] = row['energy']
        ws[f'E{index+16}'] = row['loudness']
        ws[f'F{index+16}'] = row['liveness']
        ws[f'G{index+16}'] = row['tempo']
        
    return wb

def download_excel_report(wb):
    excel_bytes = BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    return excel_bytes